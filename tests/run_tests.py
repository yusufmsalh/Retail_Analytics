"""
tests/run_tests.py
══════════════════════════════════════════════════════════════════
Self-contained test runner — no pytest required.
Tests every pure-logic function across all four features.

Run:
    python tests/run_tests.py          # in-sandbox
    pytest tests/run_tests.py -v       # locally after pip install pytest
"""

import os
import sys
import traceback
import unittest
from datetime import date, datetime, timedelta
from unittest.mock import MagicMock, patch

import numpy as np
import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

# ── Imports under test ────────────────────────────────────────────────────────
from scripts.daily_sales_report import calculate_metrics, export_to_excel, simulate_email
from scripts.inventory_forecast import (
    build_daily_demand,
    forecast_all_products,
    linear_regression_forecast,
    moving_average_forecast,
    export_forecast,
)
from scripts.low_stock_alert import detect_low_stock, send_slack_alert
from scripts.anomaly_detection import (
    detect_high_quantity_orders,
    detect_negative_stock,
    detect_sales_spikes,
    export_anomalies,
    log_summary,
)


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 1 — Daily Sales Report
# ══════════════════════════════════════════════════════════════════════════════

class TestCalculateMetrics(unittest.TestCase):
    """Tests for calculate_metrics() — the KPI engine."""

    def _make_df(self, rows):
        df = pd.DataFrame(rows, columns=[
            "order_id", "order_date", "customer_name",
            "product_id", "product_name", "sku", "quantity", "unit_price",
        ])
        df["line_total"] = df["quantity"] * df["unit_price"]
        return df

    def test_total_revenue_correct(self):
        """Revenue = sum of all line_totals."""
        df = self._make_df([
            (1, datetime.now(), "Alice", 10, "Widget A", "WA-001", 2, 10.0),
            (1, datetime.now(), "Alice", 20, "Widget B", "WB-002", 1, 20.0),
            (2, datetime.now(), "Bob",   10, "Widget A", "WA-001", 5, 10.0),
        ])
        # 2×10 + 1×20 + 5×10 = 20+20+50 = 90
        self.assertEqual(calculate_metrics(df)["total_revenue"], 90.0)

    def test_total_orders_counts_unique_order_ids(self):
        df = self._make_df([
            (1, datetime.now(), "A", 1, "P", "P-1", 1, 5.0),
            (1, datetime.now(), "A", 2, "Q", "Q-1", 2, 5.0),
            (2, datetime.now(), "B", 1, "P", "P-1", 1, 5.0),
            (3, datetime.now(), "C", 1, "P", "P-1", 3, 5.0),
        ])
        self.assertEqual(calculate_metrics(df)["total_orders"], 3)

    def test_top_products_sorted_by_revenue_descending(self):
        df = self._make_df([
            (1, datetime.now(), "A", 1, "Cheap",      "C-1", 1,  1.0),
            (2, datetime.now(), "B", 2, "Expensive",  "E-1", 1, 99.0),
            (3, datetime.now(), "C", 3, "Mid",        "M-1", 1, 10.0),
        ])
        tp = calculate_metrics(df)["top_products"]
        revenues = tp["revenue"].tolist()
        self.assertEqual(revenues, sorted(revenues, reverse=True))

    def test_top_products_max_10_rows(self):
        rows = [
            (i, datetime.now(), "X", i, f"Product {i}", f"P-{i:03d}", 1, float(i))
            for i in range(1, 20)
        ]
        df = self._make_df(rows)
        tp = calculate_metrics(df)["top_products"]
        self.assertLessEqual(len(tp), 10)

    def test_empty_dataframe_returns_zero_metrics(self):
        m = calculate_metrics(pd.DataFrame())
        self.assertEqual(m["total_revenue"], 0.0)
        self.assertEqual(m["total_orders"],  0)
        self.assertTrue(m["top_products"].empty)

    def test_single_order_single_product(self):
        df = self._make_df([(1, datetime.now(), "Z", 7, "Solo", "S-007", 4, 25.0)])
        m = calculate_metrics(df)
        self.assertEqual(m["total_revenue"], 100.0)
        self.assertEqual(m["total_orders"],  1)

    def test_revenue_rounded_to_two_decimals(self):
        df = self._make_df([(1, datetime.now(), "R", 1, "P", "P-1", 1, 0.001)])
        m = calculate_metrics(df)
        self.assertEqual(m["total_revenue"], round(0.001, 2))

    def test_top_products_contains_required_columns(self):
        df = self._make_df([(1, datetime.now(), "A", 1, "P", "P-1", 2, 5.0)])
        tp = calculate_metrics(df)["top_products"]
        for col in ["product_name", "sku", "units_sold", "revenue"]:
            self.assertIn(col, tp.columns)


class TestExportToExcel(unittest.TestCase):
    """Tests for export_to_excel() — file creation and sheet structure."""

    def setUp(self):
        os.environ["OUTPUT_DIR"] = "/tmp/test_output"
        os.makedirs("/tmp/test_output", exist_ok=True)
        self.df = pd.DataFrame({
            "order_id": [1], "order_date": [datetime.now()],
            "customer_name": ["X"], "product_id": [1],
            "product_name": ["P"], "sku": ["P-1"],
            "quantity": [2], "unit_price": [10.0], "line_total": [20.0],
        })
        self.metrics = {
            "total_revenue": 20.0,
            "total_orders":  1,
            "top_products":  pd.DataFrame({
                "product_id": [1], "product_name": ["P"], "sku": ["P-1"],
                "units_sold": [2],  "revenue": [20.0],
            }),
        }

    def test_creates_xlsx_file(self):
        path = export_to_excel(self.df, self.metrics, date(2024, 1, 15))
        self.assertTrue(os.path.exists(path))
        self.assertTrue(path.endswith(".xlsx"))

    def test_filename_contains_date(self):
        path = export_to_excel(self.df, self.metrics, date(2024, 6, 30))
        self.assertIn("20240630", os.path.basename(path))

    def test_excel_has_three_sheets(self):
        from openpyxl import load_workbook
        path = export_to_excel(self.df, self.metrics, date(2024, 1, 20))
        wb   = load_workbook(path, read_only=True)
        self.assertIn("Summary",      wb.sheetnames)
        self.assertIn("Top Products", wb.sheetnames)
        self.assertIn("Raw Orders",   wb.sheetnames)

    def test_summary_sheet_has_revenue(self):
        from openpyxl import load_workbook
        path = export_to_excel(self.df, self.metrics, date(2024, 1, 21))
        wb   = load_workbook(path, read_only=True)
        ws   = wb["Summary"]
        headers = [cell.value for cell in list(ws.rows)[0]]
        self.assertIn("Total Revenue ($)", headers)


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 2 — Inventory Forecasting
# ══════════════════════════════════════════════════════════════════════════════

class TestMovingAverageForecast(unittest.TestCase):
    """Tests for moving_average_forecast()."""

    def test_flat_series_returns_exact_value(self):
        s = pd.Series([5.0] * 20)
        self.assertAlmostEqual(moving_average_forecast(s, window=7), 5.0)

    def test_uses_only_last_window_values(self):
        # Huge outlier at the start; last 7 are all 2.0
        s = pd.Series([1000.0] + [2.0] * 10)
        self.assertAlmostEqual(moving_average_forecast(s, window=7), 2.0)

    def test_short_series_falls_back_to_full_mean(self):
        s = pd.Series([4.0, 6.0])             # shorter than window=7
        self.assertAlmostEqual(moving_average_forecast(s, window=7), 5.0)

    def test_single_value_returns_that_value(self):
        self.assertAlmostEqual(moving_average_forecast(pd.Series([42.0])), 42.0)

    def test_empty_series_returns_zero(self):
        self.assertEqual(moving_average_forecast(pd.Series([], dtype=float)), 0.0)

    def test_window_of_one_returns_last_value(self):
        s = pd.Series([1.0, 2.0, 3.0, 99.0])
        self.assertAlmostEqual(moving_average_forecast(s, window=1), 99.0)


class TestLinearRegressionForecast(unittest.TestCase):
    """Tests for linear_regression_forecast()."""

    def test_returns_correct_horizon_length(self):
        s = pd.Series(range(1, 15), dtype=float)
        self.assertEqual(len(linear_regression_forecast(s, horizon=7)), 7)

    def test_flat_series_predicts_constant(self):
        s = pd.Series([10.0] * 20)
        preds = linear_regression_forecast(s, horizon=5)
        for p in preds:
            self.assertAlmostEqual(p, 10.0, places=1)

    def test_rising_series_gives_increasing_predictions(self):
        s = pd.Series(range(1, 20), dtype=float)
        preds = linear_regression_forecast(s, horizon=3)
        self.assertLess(preds[0], preds[1])
        self.assertLess(preds[1], preds[2])

    def test_downward_trend_clipped_to_zero(self):
        s = pd.Series([100.0, 50.0, 10.0, 2.0, 0.5], dtype=float)
        preds = linear_regression_forecast(s, horizon=5)
        self.assertTrue(all(p >= 0 for p in preds))

    def test_single_value_falls_back_gracefully(self):
        preds = linear_regression_forecast(pd.Series([7.0]), horizon=4)
        self.assertEqual(len(preds), 4)
        self.assertTrue(all(p >= 0 for p in preds))

    def test_two_points_produces_valid_forecast(self):
        preds = linear_regression_forecast(pd.Series([1.0, 3.0]), horizon=3)
        self.assertEqual(len(preds), 3)


class TestBuildDailyDemand(unittest.TestCase):
    """Tests for build_daily_demand()."""

    def _base_df(self, days=10, qty=3):
        base = date(2024, 1, 1)
        return pd.DataFrame([{
            "sale_date": base + timedelta(days=i),
            "order_date": datetime(2024, 1, 1) + timedelta(days=i),
            "product_id": 1, "product_name": "Widget A",
            "sku": "WA-001", "quantity": qty,
        } for i in range(days)])

    def test_output_contains_all_products(self):
        df = pd.concat([
            self._base_df(),
            self._base_df().assign(product_id=2, product_name="Widget B", sku="WB-002"),
        ])
        result = build_daily_demand(df)
        self.assertEqual(set(result["product_id"].unique()), {1, 2})

    def test_gap_dates_filled_with_zero(self):
        """Only day 1 and day 5 have sales; days 2–4 should be zero-filled."""
        df = pd.DataFrame([
            {"sale_date": date(2024, 1, 1), "order_date": datetime(2024, 1, 1),
             "product_id": 1, "product_name": "P", "sku": "P-1", "quantity": 5},
            {"sale_date": date(2024, 1, 5), "order_date": datetime(2024, 1, 5),
             "product_id": 1, "product_name": "P", "sku": "P-1", "quantity": 3},
        ])
        result = build_daily_demand(df)
        self.assertEqual(len(result), 5)
        zeros = result[result["units_sold"] == 0]
        self.assertEqual(len(zeros), 3)   # Jan 2, 3, 4

    def test_aggregates_multiple_orders_same_day(self):
        df = pd.DataFrame([
            {"sale_date": date(2024, 1, 1), "order_date": datetime(2024, 1, 1),
             "product_id": 1, "product_name": "P", "sku": "P-1", "quantity": 4},
            {"sale_date": date(2024, 1, 1), "order_date": datetime(2024, 1, 1),
             "product_id": 1, "product_name": "P", "sku": "P-1", "quantity": 6},
        ])
        result = build_daily_demand(df)
        self.assertEqual(result.iloc[0]["units_sold"], 10)

    def test_empty_input_returns_empty(self):
        self.assertTrue(build_daily_demand(pd.DataFrame()).empty)

    def test_no_negative_units(self):
        result = build_daily_demand(self._base_df())
        self.assertTrue((result["units_sold"] >= 0).all())


class TestForecastAllProducts(unittest.TestCase):
    """Tests for forecast_all_products()."""

    def _daily_df(self, n_products=2, n_days=14):
        base  = date(2024, 1, 1)
        rows  = []
        for pid in range(1, n_products + 1):
            for i in range(n_days):
                rows.append({
                    "sale_date":    base + timedelta(days=i),
                    "product_id":   pid,
                    "product_name": f"Widget {pid}",
                    "sku":          f"W{pid:03d}",
                    "units_sold":   pid * 2,   # deterministic demand
                })
        return pd.DataFrame(rows)

    def test_correct_number_of_rows(self):
        df     = self._daily_df(n_products=3)
        result = forecast_all_products(df, horizon=7)
        # 3 products × 7 days
        self.assertEqual(len(result), 21)

    def test_recommended_reorder_non_negative(self):
        df     = self._daily_df()
        result = forecast_all_products(df, horizon=7)
        self.assertTrue((result["recommended_reorder"] >= 0).all())

    def test_all_forecast_dates_in_future(self):
        df     = self._daily_df()
        result = forecast_all_products(df, horizon=7)
        today  = date.today()
        self.assertTrue(all(d > today for d in result["forecast_date"]))

    def test_empty_input_returns_empty(self):
        self.assertTrue(forecast_all_products(pd.DataFrame(), horizon=7).empty)

    def test_recommended_reorder_is_max_of_two_models(self):
        df     = self._daily_df(n_products=1)
        result = forecast_all_products(df, horizon=7)
        for _, row in result.iterrows():
            self.assertGreaterEqual(
                row["recommended_reorder"],
                row["sma_forecast_units"],
            )
            self.assertGreaterEqual(
                row["recommended_reorder"],
                row["lr_forecast_units"],
            )


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 3 — Low Stock Alerts
# ══════════════════════════════════════════════════════════════════════════════

class TestDetectLowStock(unittest.TestCase):
    """Tests for detect_low_stock()."""

    PRODUCTS = [
        {"id": 1, "name": "Widget A", "sku": "WA-001", "stockQuantity": 3},
        {"id": 2, "name": "Widget B", "sku": "WB-002", "stockQuantity": 15},
        {"id": 3, "name": "Widget C", "sku": "WC-003", "stockQuantity": 0},
        {"id": 4, "name": "Widget D", "sku": "WD-004", "stockQuantity": 9},
        {"id": 5, "name": "Widget E", "sku": "WE-005", "stockQuantity": 50},
    ]

    def test_flags_products_below_threshold(self):
        alerts = detect_low_stock(self.PRODUCTS, threshold=10)
        ids = [a["id"] for a in alerts]
        self.assertIn(1, ids)   # qty=3
        self.assertIn(3, ids)   # qty=0
        self.assertIn(4, ids)   # qty=9

    def test_does_not_flag_products_at_or_above_threshold(self):
        alerts = detect_low_stock(self.PRODUCTS, threshold=10)
        ids = [a["id"] for a in alerts]
        self.assertNotIn(2, ids)   # qty=15
        self.assertNotIn(5, ids)   # qty=50

    def test_exact_threshold_not_flagged(self):
        """Stock == threshold should NOT trigger (strictly less-than)."""
        p = [{"id": 1, "name": "X", "sku": "X-1", "stockQuantity": 10}]
        self.assertEqual(detect_low_stock(p, threshold=10), [])

    def test_zero_stock_is_flagged(self):
        p = [{"id": 1, "name": "X", "sku": "X-1", "stockQuantity": 0}]
        self.assertEqual(len(detect_low_stock(p, threshold=1)), 1)

    def test_empty_list_returns_empty(self):
        self.assertEqual(detect_low_stock([]), [])

    def test_all_healthy_returns_empty(self):
        p = [{"id": i, "name": f"P{i}", "sku": f"P-{i}", "stockQuantity": 100}
             for i in range(5)]
        self.assertEqual(detect_low_stock(p, threshold=10), [])

    def test_alert_has_required_fields(self):
        p = [{"id": 1, "name": "X", "sku": "X-1", "stockQuantity": 3}]
        alerts = detect_low_stock(p, threshold=10)
        self.assertEqual(len(alerts), 1)
        for field in ["id", "name", "sku", "stock_quantity", "threshold", "detected_at"]:
            self.assertIn(field, alerts[0])

    def test_accepts_snake_case_field_name(self):
        """Accept stock_quantity (snake_case) as well as stockQuantity."""
        p = [{"id": 1, "name": "X", "sku": "X-1", "stock_quantity": 2}]
        self.assertEqual(len(detect_low_stock(p, threshold=10)), 1)

    def test_missing_qty_field_skipped(self):
        p = [{"id": 1, "name": "Ghost", "sku": "G-0"}]   # no qty field
        self.assertEqual(detect_low_stock(p, threshold=10), [])

    def test_custom_low_threshold(self):
        # threshold=2 means flag if qty < 2
        # Product 1 has qty=3  →  3 >= 2  →  NOT flagged
        # Product 3 has qty=0  →  0 < 2   →  flagged
        alerts = detect_low_stock(self.PRODUCTS, threshold=2)
        ids = [a["id"] for a in alerts]
        self.assertNotIn(1, ids)  # qty=3 >= 2 → not flagged
        self.assertIn(3, ids)     # qty=0 < 2  → flagged

    def test_threshold_zero_flags_nothing(self):
        """Nothing has stock < 0 except truly negative values."""
        alerts = detect_low_stock(self.PRODUCTS, threshold=0)
        self.assertEqual(len(alerts), 0)


class TestSendSlackAlert(unittest.TestCase):
    """Tests for send_slack_alert()."""

    ALERTS = [{"name": "Widget A", "sku": "WA-001", "stock_quantity": 3, "threshold": 10}]

    def test_no_http_call_for_empty_alerts(self):
        with patch("scripts.low_stock_alert.requests.post") as mock_post:
            send_slack_alert([])
            mock_post.assert_not_called()

    def test_no_http_call_when_webhook_not_configured(self):
        with patch("scripts.low_stock_alert.SLACK_WEBHOOK_URL", ""):
            with patch("scripts.low_stock_alert.requests.post") as mock_post:
                send_slack_alert(self.ALERTS)
                mock_post.assert_not_called()

    def test_http_call_made_when_real_webhook_set(self):
        real_url = "https://hooks.slack.com/services/REAL/REAL/REAL"
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        with patch("scripts.low_stock_alert.SLACK_WEBHOOK_URL", real_url):
            with patch("scripts.low_stock_alert.requests.post", return_value=mock_resp) as mp:
                send_slack_alert(self.ALERTS)
                mp.assert_called_once()
                _, kwargs = mp.call_args
                self.assertIn("blocks", kwargs["json"])

    def test_payload_contains_product_name(self):
        real_url = "https://hooks.slack.com/services/REAL/REAL/REAL"
        mock_resp = MagicMock(); mock_resp.raise_for_status = MagicMock()
        captured = {}
        def capture(*args, **kwargs):
            captured.update(kwargs)
            return mock_resp
        with patch("scripts.low_stock_alert.SLACK_WEBHOOK_URL", real_url):
            with patch("scripts.low_stock_alert.requests.post", side_effect=capture):
                send_slack_alert(self.ALERTS)
        payload_str = str(captured.get("json", ""))
        self.assertIn("Widget A", payload_str)


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 4 — Anomaly Detection
# ══════════════════════════════════════════════════════════════════════════════

def _make_order_df(n_days=30, product_id=1, base_qty=5, spike_qty=None, spike_day=None):
    """Helper: build a flat orders DataFrame, optionally injecting one spike."""
    base = datetime(2024, 1, 1)
    rows = []
    for i in range(n_days):
        qty = spike_qty if (spike_day is not None and i == spike_day) else base_qty
        rows.append({
            "order_id":    i + 1,
            "order_date":  base + timedelta(days=i),
            "product_id":  product_id,
            "product_name": f"Widget {product_id}",
            "sku":         f"W{product_id:03d}",
            "quantity":    qty,
            "unit_price":  10.0,
            "line_total":  qty * 10.0,
            "sale_date":   (base + timedelta(days=i)).date(),
        })
    return pd.DataFrame(rows)


class TestDetectHighQuantityOrders(unittest.TestCase):

    def test_flags_extreme_outlier(self):
        df     = _make_order_df(n_days=30, base_qty=5, spike_qty=500, spike_day=29)
        result = detect_high_quantity_orders(df, z_threshold=2.0)
        self.assertFalse(result.empty)
        self.assertIn(30, result["order_id"].values)  # order_id=30 is spike

    def test_uniform_data_produces_no_flags(self):
        df     = _make_order_df(n_days=30, base_qty=5)
        result = detect_high_quantity_orders(df, z_threshold=3.0)
        self.assertTrue(result.empty)

    def test_required_columns_present(self):
        df     = _make_order_df(n_days=30, base_qty=5, spike_qty=500, spike_day=15)
        result = detect_high_quantity_orders(df, z_threshold=2.0)
        for col in ["order_id", "product_name", "z_score", "anomaly_type", "detail"]:
            self.assertIn(col, result.columns)

    def test_anomaly_type_label(self):
        df     = _make_order_df(n_days=30, base_qty=5, spike_qty=999, spike_day=0)
        result = detect_high_quantity_orders(df, z_threshold=2.0)
        if not result.empty:
            self.assertTrue((result["anomaly_type"] == "High Quantity Order").all())

    def test_empty_input(self):
        self.assertTrue(detect_high_quantity_orders(pd.DataFrame()).empty)

    def test_z_scores_are_positive_for_flagged_rows(self):
        df     = _make_order_df(n_days=30, base_qty=5, spike_qty=500, spike_day=29)
        result = detect_high_quantity_orders(df, z_threshold=2.0)
        self.assertTrue((result["z_score"] > 0).all())

    def test_higher_threshold_flags_fewer_rows(self):
        df    = _make_order_df(n_days=30, base_qty=5, spike_qty=500, spike_day=29)
        low_t = detect_high_quantity_orders(df, z_threshold=1.0)
        hi_t  = detect_high_quantity_orders(df, z_threshold=5.0)
        self.assertGreaterEqual(len(low_t), len(hi_t))


class TestDetectNegativeStock(unittest.TestCase):

    PRODUCTS = pd.DataFrame({
        "product_id":    [1, 2, 3, 4],
        "product_name":  ["A", "B", "C", "D"],
        "sku":           ["A-1", "B-2", "C-3", "D-4"],
        "stock_quantity": [50, -5, 0, -1],
    })

    def test_flags_negative_stock(self):
        result = detect_negative_stock(self.PRODUCTS)
        ids    = set(result["product_id"].tolist())
        self.assertIn(2, ids)   # -5
        self.assertIn(4, ids)   # -1

    def test_zero_stock_not_flagged(self):
        result = detect_negative_stock(self.PRODUCTS)
        self.assertNotIn(3, result["product_id"].tolist())

    def test_positive_stock_not_flagged(self):
        result = detect_negative_stock(self.PRODUCTS)
        self.assertNotIn(1, result["product_id"].tolist())

    def test_correct_anomaly_type(self):
        result = detect_negative_stock(self.PRODUCTS)
        self.assertTrue((result["anomaly_type"] == "Negative Stock").all())

    def test_all_positive_returns_empty(self):
        df = pd.DataFrame({"product_id": [1], "product_name": ["X"],
                           "sku": ["X-1"], "stock_quantity": [10]})
        self.assertTrue(detect_negative_stock(df).empty)

    def test_empty_input_returns_empty(self):
        self.assertTrue(detect_negative_stock(pd.DataFrame()).empty)

    def test_detected_at_field_present(self):
        result = detect_negative_stock(self.PRODUCTS)
        self.assertIn("detected_at", result.columns)


class TestDetectSalesSpikes(unittest.TestCase):

    def _spike_df(self, n_normal=29, spike_revenue=10000.0):
        base = date(2024, 1, 1)
        rows = [
            {"sale_date": base + timedelta(days=i), "product_id": 1,
             "product_name": "P", "sku": "P-1", "line_total": 100.0,
             "order_date": datetime(2024, 1, 1) + timedelta(days=i),
             "quantity": 1, "unit_price": 100.0}
            for i in range(n_normal)
        ]
        # Add spike on last day
        rows.append({
            "sale_date": base + timedelta(days=n_normal), "product_id": 1,
            "product_name": "P", "sku": "P-1", "line_total": spike_revenue,
            "order_date": datetime(2024, 1, 1) + timedelta(days=n_normal),
            "quantity": 100, "unit_price": 100.0,
        })
        return pd.DataFrame(rows)

    def test_flags_large_revenue_spike(self):
        df     = self._spike_df()
        result = detect_sales_spikes(df, z_threshold=2.0)
        self.assertFalse(result.empty)

    def test_uniform_revenue_produces_no_flags(self):
        df     = _make_order_df(n_days=30, base_qty=5)
        result = detect_sales_spikes(df, z_threshold=3.0)
        self.assertTrue(result.empty)

    def test_required_columns_present(self):
        df     = self._spike_df()
        result = detect_sales_spikes(df, z_threshold=2.0)
        for col in ["sale_date", "product_name", "z_score", "anomaly_type", "detail"]:
            self.assertIn(col, result.columns)

    def test_anomaly_type_label(self):
        df     = self._spike_df()
        result = detect_sales_spikes(df, z_threshold=2.0)
        if not result.empty:
            self.assertTrue((result["anomaly_type"] == "Sales Spike").all())

    def test_empty_input_returns_empty(self):
        self.assertTrue(detect_sales_spikes(pd.DataFrame()).empty)

    def test_spike_date_is_last_day(self):
        df     = self._spike_df(n_normal=29)
        result = detect_sales_spikes(df, z_threshold=2.0)
        if not result.empty:
            spike_date = date(2024, 1, 1) + timedelta(days=29)
            self.assertIn(spike_date, result["sale_date"].tolist())


class TestExportAnomalies(unittest.TestCase):

    def setUp(self):
        os.environ["OUTPUT_DIR"] = "/tmp/test_anomaly_output"

    def _anomalies(self, include_data=True):
        df = pd.DataFrame({
            "product_id": [1], "product_name": ["P"], "sku": ["P-1"],
            "stock_quantity": [-3], "anomaly_type": ["Negative Stock"],
            "detail": ["stock=-3"], "detected_at": [datetime.now().isoformat()],
        }) if include_data else pd.DataFrame()
        return {
            "High Quantity Orders": pd.DataFrame(),
            "Negative Stock":       df,
            "Sales Spikes":         pd.DataFrame(),
        }

    def test_creates_xlsx_file(self):
        path = export_anomalies(self._anomalies())
        self.assertTrue(os.path.exists(path))
        self.assertTrue(path.endswith(".xlsx"))

    def test_excel_has_three_sheets(self):
        from openpyxl import load_workbook
        path = export_anomalies(self._anomalies())
        wb   = load_workbook(path, read_only=True)
        for sheet in ["High Quantity Orders", "Negative Stock", "Sales Spikes"]:
            self.assertIn(sheet, wb.sheetnames)

    def test_empty_anomalies_still_creates_file(self):
        path = export_anomalies(self._anomalies(include_data=False))
        self.assertTrue(os.path.exists(path))


# ══════════════════════════════════════════════════════════════════════════════
# RUNNER
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    loader  = unittest.TestLoader()
    suite   = unittest.TestSuite()

    test_classes = [
        # Feature 1
        TestCalculateMetrics,
        TestExportToExcel,
        # Feature 2
        TestMovingAverageForecast,
        TestLinearRegressionForecast,
        TestBuildDailyDemand,
        TestForecastAllProducts,
        # Feature 3
        TestDetectLowStock,
        TestSendSlackAlert,
        # Feature 4
        TestDetectHighQuantityOrders,
        TestDetectNegativeStock,
        TestDetectSalesSpikes,
        TestExportAnomalies,
    ]

    for cls in test_classes:
        suite.addTests(loader.loadTestsFromTestCase(cls))

    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    # Summary
    total  = result.testsRun
    passed = total - len(result.failures) - len(result.errors)
    print("\n" + "═" * 60)
    print(f"  Tests run:  {total}")
    print(f"  Passed:     {passed}")
    print(f"  Failed:     {len(result.failures)}")
    print(f"  Errors:     {len(result.errors)}")
    print("═" * 60)

    sys.exit(0 if result.wasSuccessful() else 1)
